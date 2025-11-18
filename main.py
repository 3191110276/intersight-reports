#!/usr/bin/env python
# coding: utf-8

"""
This tool allows you to export data from Cisco Intersight into .xlsx file reports.
"""

# ----------------------------------------------------------------------
# Import Dependencies
# ----------------------------------------------------------------------
# With any programming language, we will need to import dependencies unless we want to build everything from scratch.
# We will briefly go over the dependencies that need to be imported and what they will be used for.
# The details of their usage will be discussed later on when we use the dependencies to build some functionality with them.

# The *argparse* module is used to provide a minimal command line when running the script
import argparse

# The *logging* module will allow us to add log messages to understand what is going on inside of the application
import logging

# The *os* module will be used for various file operations
import os

# The *re* module will be used to help with authentication
import re

# The *time* module is used to keep the application alive
import time

# The *datetime* and *pytz* modules are used to get the current time and to create timestamps
from datetime import datetime, timedelta
import pytz

# The *intersight* modules are used to handle the authentication to Intersight
import intersight

# We import the telemetry API object from the *intersight* module to build a metrics request
from intersight.api import telemetry_api

# We import the *yaml* module, which we will be using to read/write configurations and metadata for our tool
import yaml

# We use openpyxl for handling xlsx files
from openpyxl import load_workbook

# ----------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------
# To allow for changing the application settings, we need to get the configuration from somewhere
# In this tool, we read configuration from a YAML configuration file
# This is a very simple way of specifying configuration, as it is easily readable
# The functions below are used to read and parse the configuration file
def load_config(config_file_path=None):
    '''Reads and parses the configuration'''
    if config_file_path is None:
        config_file_path = 'config.yaml'

    if not os.path.isabs(config_file_path):
        config_file_path = os.path.abspath(config_file_path)

    with open(config_file_path, 'r', encoding="utf-8") as file:
        parsed_config = yaml.safe_load(file)

    return parsed_config


def safe_load(base, value, default):
    '''Loads a value from the provided structure, or returns the default value if none is found'''
    return base[value] if value in base else default

# ----------------------------------------------------------------------
# Intersight Authentication
# ----------------------------------------------------------------------
# To start running requests to Intersight, we first need to create an API client for authentication
def get_api_client(api_key_id, api_secret_path, endpoint="https://intersight.com"):
    '''
    Creates an Intersight API client that handles authentication
    This is based on the authentication example from the official GitHub repository of the Intersight Python module
    You can find the repository here: https://github.com/CiscoDevNet/intersight-python
    '''
    if not os.path.exists(api_secret_path):
        logging.critical('Secret Key file does not exist!')

    with open(api_secret_path, 'r', encoding="utf-8") as api_secret_file:
        api_key = api_secret_file.read()
        if api_key == '':
            logging.critical('Secret Key file is empty!')

    signing_alg = None
    signing_scheme = None
    hash_alg = None

    if re.search('BEGIN RSA PRIVATE KEY', api_key):
        # API Key v2 format
        logger.info('Using Intersight v2 API Key')
        signing_alg = intersight.signing.ALGORITHM_RSASSA_PKCS1v15
        signing_scheme = intersight.signing.SCHEME_RSA_SHA256
        hash_alg = intersight.signing.HASH_SHA256

    elif re.search('BEGIN EC PRIVATE KEY', api_key):
        # API Key v3 format
        logger.info('Using Intersight v3 API Key')
        signing_alg = intersight.signing.ALGORITHM_ECDSA_MODE_DETERMINISTIC_RFC6979
        signing_scheme = intersight.signing.SCHEME_HS2019
        hash_alg = intersight.signing.HASH_SHA256

    configuration = intersight.Configuration(
        host=endpoint,
        signing_info=intersight.signing.HttpSigningConfiguration(
            key_id=api_key_id,
            private_key_path=api_secret_path,
            signing_scheme=signing_scheme,
            signing_algorithm=signing_alg,
            hash_algorithm=hash_alg,
            signed_headers=[
                intersight.signing.HEADER_REQUEST_TARGET,
                intersight.signing.HEADER_HOST,
                intersight.signing.HEADER_DATE,
                intersight.signing.HEADER_DIGEST,
            ]
        )
    )

    return intersight.ApiClient(configuration)

# ----------------------------------------------------------------------
# Retrieval function
# ----------------------------------------------------------------------
# This function will execute our query, and then parse the response into a list of rows
# First, we need to define the function that retrieves the data
# This will use the built-in functions of the Python SDK to call the correct API endpoint
# We are passing our request that we defined before as an input. It will be used as the body of our request
# Everything else uses default values or presets that we are not going to explain in detail
# We do not need to ouch any of the other settings
# Once the request returns a result, it will call our function to parse the API output into something that is useful for us
def perform_query(api_client, req, offset=0):
    '''Performs the HTTP query to Intersight to fetch data'''
    logger.debug('Performing query for request: %s', req)

    try:
        if 'limitSpec' in req:
            req['limitSpec']['offset'] = offset
        api_response = api_client.api_client.call_api(
            resource_path=api_client.query_telemetry_scan_endpoint.settings["endpoint_path"],
            method='POST',
            header_params={'Accept': 'application/json', 'Content-Type': 'application/json'},
            body=req,
            _host=IS_ENDPOINT,
            async_req=False,
            auth_settings=api_client.query_telemetry_scan_endpoint.settings['auth'],
            response_type=api_client.query_telemetry_scan_endpoint.settings['response_type'],
            _check_type=True,
            _return_http_data_only=True,
            _preload_content=False
        )

        output = parse_response(api_response.json(), req)

        if len(output['row_output']) > 0:
            if len(output['row_output']) == req['limitSpec']['limit']:
                extra_output = perform_query(api_client, req, req['limitSpec']['offset'] + req['limitSpec']['limit'])
                output['row_output'].extend(extra_output['data']['row_output'])
            return {
                'status': 'success',
                'data': output
            }

        return {
            'status': 'empty',
            'data': output
        }

    except intersight.exceptions.ApiException as e:
        if 'Druid response exceeds maximum response size' in str(e):
            logger.error('The response size exceeds the maxiumum allowed by the Intersight API!')
            logger.error('Lower the >limit< variable in your configuration file to receive data from Intersight.')
            return {
                'status': 'size_exceeded',
                'data': []
            }

        logger.error('Query failed: %s', e)
        return {
            'status': 'fail',
            'data': []
        }


def connectivity_test(api_client, req):
    '''Tests the connectivity to Intersight to ensure that requests can succeed'''
    try:
        api_response = api_client.api_client.call_api(
            resource_path=api_client.query_telemetry_scan_endpoint.settings["endpoint_path"],
            method='POST',
            header_params={'Accept': 'application/json', 'Content-Type': 'application/json'},
            body=req,
            _host=IS_ENDPOINT,
            async_req=False,
            auth_settings=api_client.query_telemetry_scan_endpoint.settings['auth'],
            response_type=api_client.query_telemetry_scan_endpoint.settings['response_type'],
            _check_type=False,
            _return_http_data_only=True,
            _preload_content=False
        )
        if 'timestamp' not in api_response.json()[0]:
            raise ConnectionAbortedError
        logger.info('Intersight connectivity check succeeded')
    except Exception as e:
        logger.error('Connectivity to Intersight failed: %s', e)


def parse_response(res, req):
    '''Transforms the response of an Intersight metrics Scan request into a useable output'''
    keys = []
    row_output = []

    for key in res[0]['event']:
        keys.append(key)

    if len(res) > 0:
        for row in res:
            event_out = {
                'timestamp': row['timestamp']
            }
            for key in keys:
                event_out[key] = row['event'][key]

            row_output.append(event_out)

    keys.insert(0, "timestamp")

    return {'keys': keys, 'row_output': row_output}


# ----------------------------------------------------------------------
# Creating Timestamps
# ----------------------------------------------------------------------
# Timestamps have to be created for each request time range
# The timestamps have to follow the ISO-8601 format
def create_interval():
    '''
    Creates a time range that can be used as a time boundary in Intersight requests
    Timestamps are rounded to the closest minute, as Intersight does not support resolutions below a minute anyways
    There is an offset from the current time to ensure that data that is very recent is not queried right away
    Data that was just collected from an endpoint might not be visible by users yet
    '''

    current_time_utc = datetime.now().astimezone(pytz.utc)
    rounded = current_time_utc.replace(minute=0, second=0, microsecond=0)

    start = (rounded - timedelta(days=365)).isoformat().replace("+00:00", "Z")
    end = rounded.isoformat().replace("+00:00", "Z")

    return f"{start}/{end}"

# ----------------------------------------------------------------------
# Excel Data Writer
# ----------------------------------------------------------------------
def write_excel_output_data(ws, new_data):
    excel_headers = [cell.value for cell in ws[1] if cell.value]

    # Clear old data (except headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(excel_headers)):
        for cell in row:
            cell.value = None

    # Insert new data dynamically
    for i, record in enumerate(new_data, start=2):
        for j, header in enumerate(excel_headers, start=1):
            if header in record:
                ws.cell(row=i, column=j, value=record[header])

# ----------------------------------------------------------------------
# Running the application
# ----------------------------------------------------------------------
# If this file is run, the code below will be executed
if __name__ == '__main__':
    # When running the tool, we want to provide a minimal command line interface
    # Right now this is only used to provide a path to the configuration file
    parser = argparse.ArgumentParser(description="Intersight Report Builder Tool")
    parser.add_argument("--config")
    args, leftovers = parser.parse_known_args()

    # We are loading the configuration from the config file
    config = load_config(args.config)

    # Before we start implementing our export functionality, we will configure logging
    # Having messages in a log file allows us to reproduce what was going on inside of the application in the past
    # We can look back in time to see if there have been any errors, and potentially even what their cause was
    LOG_DIRECTORY = './log'
    LOG_LEVEL = safe_load(config['logging'], 'level', 'INFO')

    if not os.path.exists(LOG_DIRECTORY):
        os.makedirs(LOG_DIRECTORY)

    logger = logging.getLogger('is_report')
    logging.basicConfig(
        filename=f"{LOG_DIRECTORY}/intersight_report.log",
        level=logging.getLevelName(LOG_LEVEL),
        format='%(asctime)s.%(msecs)03d %(levelname)s {%(module)s} [%(funcName)s] %(message)s',
        datefmt='%Y-%m-%d,%H:%M:%S'
    )

    logger.info('--------------------------------------------------------------------')
    logger.info('Intersight Report builder started')
    logger.info('--------------------------------------------------------------------')

    # Now we need to load some basic configuration from the config file
    IS_ENDPOINT = config['intersight']['endpoint']
    API_KEY = config['intersight']['api_key']
    SECRET_KEY = config['intersight']['secret_key_path']

    logger.info('Endpoint: %s', IS_ENDPOINT)
    logger.info('Log Level: %s', LOG_LEVEL)

    # To authenticate with the Intersight API, we need to create an API client
    # Because we want to create telemetry requests, we need to initiate an instance for that particular API group
    api_instance = telemetry_api.TelemetryApi(get_api_client(API_KEY, SECRET_KEY, IS_ENDPOINT))

    # Before starting the application, we will perform a connectivity test to Intersight
    # This will ensure that requests will succeed - otherwise you might wait a long time for the first request to fail
    connectivity_test(api_instance, 
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": "all",
            "limitSpec": {
                "type"    : "default",
                "limit"   : 1,
                "offset"  : 0
            },
            "intervals": ["2022-02-02/2022-02-04"],
            "dimensions": [],
            "aggregations": []
        })

    psu_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 28000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name",
            "name",
            "host.type"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.power_supply"
                }
            ]
            },
            "aggregations": [
                {
                    "type": "doubleMax",
                    "name": "power_max",
                    "fieldName": "hw.power_max"
                },
                {
                    "type": "longSum",
                    "name": "count",
                    "fieldName": "hw.power_count"
                },
                {
                    "type": "doubleSum",
                    "name": "hw.power-Sum",
                    "fieldName": "hw.power"
                },
                {
                    "type": "longLast",
                    "name": "status",
                    "fieldName": "hw.status"
                }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "power_avg",
                "expression": "(\"hw.power-Sum\" / \"count\")"
            }
            ]
        })
    
    temperature_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 40000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name",
            "parent.name",
            "sensor_location",
            "name"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.temperature"
                }
            ]
            },
            "aggregations": [
                {
                    "type": "doubleMax",
                    "name": "temeprature_max",
                    "fieldName": "hw.temperature_max"
                },
                {
                    "type": "longSum",
                    "name": "count",
                    "fieldName": "hw.temperature_count"
                },
                {
                    "type": "doubleSum",
                    "name": "hw.temperature-Sum",
                    "fieldName": "hw.temperature"
                }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "temperature_avg",
                "expression": "(\"hw.temperature-Sum\" / \"count\")"
            }
            ]
        })
    
    fanspeed_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 50000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name",
            "name",
            "hw.fan.airflow_direction"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.fan"
                }
            ]
            },
            "aggregations": [
                {
                    "type": "longMax",
                    "name": "fanspeed_max",
                    "fieldName": "hw.fan.speed_max"
                },
                {
                    "type": "longSum",
                    "name": "count",
                    "fieldName": "hw.fan.speed_count"
                },
                {
                    "type": "longSum",
                    "name": "hw.fan.speed-Sum",
                    "fieldName": "hw.fan.speed"
                }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "fanspeed_avg",
                "expression": "(\"hw.fan.speed-Sum\" / \"count\")"
            }
            ]
        })
    
    ecc_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1Y",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 50000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.memory"
                }
            ]
            },
            "aggregations": [
                {
                    "type": "longSum",
                    "name": "correctable_ecc",
                    "fieldName": "hw.errors_correctable_ecc_errors"
                },
                {
                    "type": "doubleSum",
                    "name": "uncorrectable_ecc",
                    "fieldName": "hw.errors_uncorrectable_ecc_errors"
                }
            ]
        })
    
    cpu_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 50000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.cpu"
                }
            ]
            },
            "aggregations": [
                {
                    "type": "doubleMax",
                    "name": "cpuutil_max",
                    "fieldName": "hw.cpu.utilization_c0_max"
                },
                {
                    "type": "longSum",
                    "name": "count",
                    "fieldName": "hw.cpu.utilization_c0_count"
                },
                {
                    "type": "doubleSum",
                    "name": "hw.cpu.utilization_c0-Sum",
                    "fieldName": "hw.cpu.utilization_c0"
                }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "cpuutil_avg",
                "expression": "(\"hw.cpu.utilization_c0-Sum\" / \"count\")"
            }
            ]
        })
    
    host_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "PhysicalEntities",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 40000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.host"
                }
            ]
            },
            "aggregations": [
            {
                "type": "doubleMax",
                "name": "hostpower_max",
                "fieldName": "hw.host.power_max"
            },
            {
                "type": "longMax",
                "name": "hoststate_max",
                "fieldName": "hw.host.power_state_max"
            },
            {
                "type": "longSum",
                "name": "power_count",
                "fieldName": "hw.host.power_count"
            },
            {
                "type": "doubleSum",
                "name": "Power-Sum",
                "fieldName": "hw.host.power"
            },
            {
                "type": "longSum",
                "name": "state_count",
                "fieldName": "hw.host.power_state_count"
            },
            {
                "type": "doubleSum",
                "name": "State-Sum",
                "fieldName": "hw.host.power_state"
            }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "hostpower_avg",
                "expression": "(\"Power-Sum\" / \"power_count\")"
            },
            {
                "type": "expression",
                "name": "hoststate_avg",
                "expression": "(\"State-Sum\" / \"state_count\")"
            }
            ]
        })
    
    network_data = perform_query(api_instance,
        {
            "queryType": "groupBy",
            "dataSource": "NetworkInterfaces",
            "granularity": {
            "type": "period",
            "period": "P1D",
            "timeZone": "Europe/Vienna"
            },
            "limitSpec": {
                "type"    : "default",
                "limit"   : 15000,
                "offset"  : 0
            },
            "intervals": [
            create_interval()
            ],
            "dimensions": [
            "intersight.domain.name",
            "host.name",
            "name",
            "host.type",
            "hw.network.port.type",
            "hw.network.port.role",
            "hw.network.port.port_channel",
            "hw.chassis.number"
            ],
            "filter": {
            "type": "and",
            "fields": [
                {
                "type": "selector",
                "dimension": "instrument.name",
                "value": "hw.network"
                }
            ]
            },
            "aggregations": [
            {
                "type": "doubleMax",
                "name": "tx_net_util_max",
                "fieldName": "hw.network.bandwidth.utilization_transmit_max"
            },
            {
                "type": "doubleMax",
                "name": "rx_net_util_max",
                "fieldName": "hw.network.bandwidth.utilization_receive_max"
            },
            {
                "type": "doubleMax",
                "name": "tx_net_bw_max",
                "fieldName": "hw.network.io_transmit_max"
            },
            {
                "type": "doubleMax",
                "name": "rx_net_bw_max",
                "fieldName": "hw.network.io_receive_max"
            },
            {
                "type": "longSum",
                "name": "transmit_util_count",
                "fieldName": "hw.network.bandwidth.utilization_transmit_count"
            },
            {
                "type": "doubleSum",
                "name": "transmit_util_sum",
                "fieldName": "hw.network.bandwidth.utilization_transmit"
            },
            {
                "type": "longSum",
                "name": "receive_util_count",
                "fieldName": "hw.network.bandwidth.utilization_receive_count"
            },
            {
                "type": "doubleSum",
                "name": "receive_util_sum",
                "fieldName": "hw.network.bandwidth.utilization_receive"
            },
            {
                "type": "doubleSum",
                "name": "transmit_bw_count",
                "fieldName": "hw.network.io_transmit_duration"
            },
            {
                "type": "longSum",
                "name": "transmit_bw_sum",
                "fieldName": "hw.network.io_receive"
            },
            {
                "type": "doubleSum",
                "name": "receive_bw_count",
                "fieldName": "hw.network.io_receive_duration"
            },
            {
                "type": "longSum",
                "name": "receive_bw_sum",
                "fieldName": "hw.network.io_receive"
            },
            {
                "type": "longLast",
                "name": "link_bw",
                "fieldName": "hw.network.bandwidth.limit"
            },
            {
                "type": "longSum",
                "name": "transmit_eth_pause",
                "fieldName": "hw.errors_network_transmit_pause"
            },
            {
                "type": "longSum",
                "name": "transmit_fc_pause",
                "fieldName": "hw.network.packets_transmit_ppp"
            },
            {
                "type": "longSum",
                "name": "receive_eth_pause",
                "fieldName": "hw.errors_network_receive_pause"
            },
            {
                "type": "longSum",
                "name": "receive_fc_pause",
                "fieldName": "hw.network.packets_receive_ppp"
            },
            {
                "type": "longSum",
                "name": "receive_crc_errors",
                "fieldName": "hw.errors_network_receive_crc"
            }
            ],
            "postAggregations": [
            {
                "type": "expression",
                "name": "tx_net_util_avg",
                "expression": "(\"transmit_util_sum\" / \"transmit_util_count\")"
            },
            {
                "type": "expression",
                "name": "rx_net_util_avg",
                "expression": "(\"receive_util_sum\" / \"receive_util_count\")"
            },
            {
                "type": "expression",
                "name": "tx_net_bw_avg",
                "expression": "(\"transmit_bw_sum\" / \"transmit_bw_count\")"
            },
            {
                "type": "expression",
                "name": "rx_net_bw_avg",
                "expression": "(\"receive_bw_sum\" / \"receive_bw_count\")"
            }
            ]
        })

    # Load your existing template
    wb = load_workbook("template.xlsx")

    # ---------------------------
    # PSU DATA
    # ---------------------------
    write_excel_output_data(wb["Data - PSUs"], psu_data['data']['row_output'])

    # ---------------------------
    # TEMPERATURE DATA
    # ---------------------------
    write_excel_output_data(wb["Data - Temperature"], temperature_data['data']['row_output'])

    # ---------------------------
    # FAN SPEED DATA
    # ---------------------------
    write_excel_output_data(wb["Data - Fan Speed"], fanspeed_data['data']['row_output'])

    # ---------------------------
    # ECC DATA
    # ---------------------------
    write_excel_output_data(wb["Data - ECC"], ecc_data['data']['row_output'])

    # ---------------------------
    # CPU DATA
    # ---------------------------
    write_excel_output_data(wb["Data - CPU"], cpu_data['data']['row_output'])

    # ---------------------------
    # HOST DATA
    # ---------------------------
    write_excel_output_data(wb["Data - Host"], host_data['data']['row_output'])

    # ---------------------------
    # NETWORK DATA
    # ---------------------------
    write_excel_output_data(wb["Data - Network"], network_data['data']['row_output'])

    # ---------------------------
    # PSU ANALYSIS
    # ---------------------------
    power_consumption_view_output = []
    hosts = {}
    for i in psu_data['data']['row_output']:
        if i['host.name'] not in hosts:
            hosts[i['host.name']] = {
                'domain_name': i['intersight.domain.name'],
                'host_type': i['host.type'],
                'power_avg': {},
                'power_max': {}
            }

        if i['power_max'] != None:
            if i['name'] not in hosts[i['host.name']]['power_max']:
                hosts[i['host.name']]['power_max'][i['name']] = 0
            if i['power_max'] > hosts[i['host.name']]['power_max'][i['name']]:
                hosts[i['host.name']]['power_max'][i['name']] = i['power_max']

        if i['power_avg'] != None:
            if i['name'] not in hosts[i['host.name']]['power_avg']:
                hosts[i['host.name']]['power_avg'][i['name']] = {'sum': 0, 'count': 0}
            hosts[i['host.name']]['power_avg'][i['name']]['sum'] += i['power_avg']
            hosts[i['host.name']]['power_avg'][i['name']]['count'] += 1

    for host in hosts:
        power_avg = 0
        for psu in hosts[host]['power_avg']:
            power_avg += hosts[host]['power_avg'][psu]['sum']/hosts[host]['power_avg'][psu]['count']

        power_max = 0
        for psu in hosts[host]['power_max']:
            power_max += hosts[host]['power_max'][psu]

        power_consumption_view_output.append([
            hosts[host]['domain_name'],
            host,
            hosts[host]['host_type'],
            f"=ROUND({power_avg},1)",
            f"=ROUND({power_max},1)"
        ])
        
    ws = wb["Chassis Power Consumption"]
    for r in power_consumption_view_output:
        ws.append(r)

    # Save to new file
    wb.save("report.xlsx")
    print("✅ report.xlsx updated successfully.")